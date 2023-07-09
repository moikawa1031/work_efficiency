import openpyxl


wb = openpyxl.load_workbook("C:\\work\\CB_VMList.xlsx")

ws = wb.active
ws["A1"] = "Nakamura"
wb.save("C:\\work\\CB_VMList_save.xlsx")
print(ws["A1"].value)